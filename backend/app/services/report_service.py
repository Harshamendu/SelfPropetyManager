import uuid
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.expense import Expense
from app.models.property import Property
from app.models.rental_payment import RentalPayment

EXPENSE_CATEGORIES = [
    "HOA",
    "Maintenance",
    "Insurance",
    "Tax",
    "Mortgage",
    "Utility",
    "Repair",
    "Other",
]

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


async def generate_year_end_excel(
    db: AsyncSession,
    year: int,
    property_id: uuid.UUID | None = None,
) -> BytesIO:
    """Generate a year-end Excel report. One sheet per property."""
    # Determine which properties to include
    prop_stmt = select(Property).where(Property.is_active == True)  # noqa: E712
    if property_id is not None:
        prop_stmt = prop_stmt.where(Property.id == property_id)
    prop_stmt = prop_stmt.order_by(Property.name)
    prop_result = await db.execute(prop_stmt)
    properties = list(prop_result.scalars().all())

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    summary_font = Font(bold=True, size=11)

    for prop in properties:
        ws = wb.create_sheet(title=prop.name[:31])  # Excel sheet name max 31 chars

        # Header row
        headers = ["Month", "Rent Collected"] + EXPENSE_CATEGORIES + ["Total Expenses"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Fetch rent data by month
        rent_stmt = (
            select(
                extract("month", RentalPayment.payment_date).label("month"),
                func.coalesce(func.sum(RentalPayment.amount), 0).label("total"),
            )
            .where(RentalPayment.property_id == prop.id)
            .where(extract("year", RentalPayment.payment_date) == year)
            .group_by(extract("month", RentalPayment.payment_date))
        )
        rent_result = await db.execute(rent_stmt)
        rent_by_month: dict[int, Decimal] = {}
        for row in rent_result:
            rent_by_month[int(row.month)] = row.total

        # Fetch expense data by month and category
        expense_stmt = (
            select(
                extract("month", Expense.date).label("month"),
                Expense.category,
                func.coalesce(func.sum(Expense.amount), 0).label("total"),
            )
            .where(Expense.property_id == prop.id)
            .where(extract("year", Expense.date) == year)
            .group_by(extract("month", Expense.date), Expense.category)
        )
        expense_result = await db.execute(expense_stmt)
        # expenses_by_month[month][category] = total
        expenses_by_month: dict[int, dict[str, Decimal]] = {}
        for row in expense_result:
            m = int(row.month)
            if m not in expenses_by_month:
                expenses_by_month[m] = {}
            expenses_by_month[m][row.category] = row.total

        # Write data rows (one per month)
        total_rent = Decimal("0")
        total_expenses_all = Decimal("0")
        category_totals: dict[str, Decimal] = {cat: Decimal("0") for cat in EXPENSE_CATEGORIES}

        for month_idx in range(1, 13):
            row_num = month_idx + 1
            ws.cell(row=row_num, column=1, value=MONTH_NAMES[month_idx - 1])

            rent = rent_by_month.get(month_idx, Decimal("0"))
            total_rent += rent
            ws.cell(row=row_num, column=2, value=float(rent))

            month_total_expenses = Decimal("0")
            month_expenses = expenses_by_month.get(month_idx, {})
            for cat_idx, cat in enumerate(EXPENSE_CATEGORIES, start=3):
                cat_amount = month_expenses.get(cat, Decimal("0"))
                month_total_expenses += cat_amount
                category_totals[cat] += cat_amount
                ws.cell(row=row_num, column=cat_idx, value=float(cat_amount))

            total_expenses_all += month_total_expenses
            ws.cell(
                row=row_num,
                column=len(EXPENSE_CATEGORIES) + 3,
                value=float(month_total_expenses),
            )

        # Summary row
        summary_row = 14
        ws.cell(row=summary_row, column=1, value="TOTAL").font = summary_font
        ws.cell(row=summary_row, column=2, value=float(total_rent)).font = summary_font
        for cat_idx, cat in enumerate(EXPENSE_CATEGORIES, start=3):
            ws.cell(
                row=summary_row, column=cat_idx, value=float(category_totals[cat])
            ).font = summary_font
        ws.cell(
            row=summary_row,
            column=len(EXPENSE_CATEGORIES) + 3,
            value=float(total_expenses_all),
        ).font = summary_font

        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 16

    # If no properties, create a placeholder sheet
    if not properties:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No properties found for the given criteria.")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


async def get_year_summary(db: AsyncSession, year: int) -> dict:
    """Return totals per property for a given year."""
    prop_stmt = (
        select(Property)
        .where(Property.is_active == True)  # noqa: E712
        .order_by(Property.name)
    )
    prop_result = await db.execute(prop_stmt)
    properties = list(prop_result.scalars().all())

    result_properties = []
    for prop in properties:
        # Total rent
        rent_stmt = (
            select(func.coalesce(func.sum(RentalPayment.amount), 0))
            .where(RentalPayment.property_id == prop.id)
            .where(extract("year", RentalPayment.payment_date) == year)
        )
        rent_result = await db.execute(rent_stmt)
        total_rent = rent_result.scalar_one()

        # Total expenses
        expense_stmt = (
            select(func.coalesce(func.sum(Expense.amount), 0))
            .where(Expense.property_id == prop.id)
            .where(extract("year", Expense.date) == year)
        )
        expense_result = await db.execute(expense_stmt)
        total_expenses = expense_result.scalar_one()

        # Monthly summaries
        monthly_summaries = []
        for month in range(1, 13):
            m_rent_stmt = (
                select(func.coalesce(func.sum(RentalPayment.amount), 0))
                .where(RentalPayment.property_id == prop.id)
                .where(extract("year", RentalPayment.payment_date) == year)
                .where(extract("month", RentalPayment.payment_date) == month)
            )
            m_rent_result = await db.execute(m_rent_stmt)
            m_rent = m_rent_result.scalar_one()

            m_exp_stmt = (
                select(func.coalesce(func.sum(Expense.amount), 0))
                .where(Expense.property_id == prop.id)
                .where(extract("year", Expense.date) == year)
                .where(extract("month", Expense.date) == month)
            )
            m_exp_result = await db.execute(m_exp_stmt)
            m_exp = m_exp_result.scalar_one()

            monthly_summaries.append({
                "month": month,
                "rent_collected": float(m_rent),
                "total_expenses": float(m_exp),
            })

        result_properties.append({
            "property_id": str(prop.id),
            "property_name": prop.name,
            "monthly_summaries": monthly_summaries,
            "total_rent": float(total_rent),
            "total_expenses": float(total_expenses),
        })

    return {
        "year": year,
        "properties": result_properties,
    }
