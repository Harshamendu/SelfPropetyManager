import io
from collections import defaultdict
from datetime import date
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import extract, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import (
    get_accessible_property_ids,
    get_db,
    require_property_access,
    require_user,
)
from app.models.expense import Expense
from app.models.property import Property
from app.models.rental_payment import RentalPayment
from app.schemas.report import (
    CategoryMonthlySummary,
    MonthlySummary,
    PropertyYearSummary,
    YearEndSummary,
)

router = APIRouter(tags=["Reports"], dependencies=[Depends(require_user)])

EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

MONTH_NAMES = [
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


def _confirmed_filter():
    today = date.today()
    return or_(
        RentalPayment.is_marked_done == True,  # noqa: E712
        RentalPayment.payment_date <= today,
    )


def _confirmed_expense_filter():
    today = date.today()
    return or_(
        Expense.is_marked_done == True,  # noqa: E712
        Expense.date <= today,
    )


async def _build_property_summary(
    db: AsyncSession,
    prop: Property,
    year: int,
) -> PropertyYearSummary:
    # --- Income by category per month ---
    income_rows = await db.execute(
        select(
            extract("month", RentalPayment.payment_date).label("month"),
            func.coalesce(RentalPayment.category, "Rent").label("category"),
            func.sum(RentalPayment.amount).label("total"),
        )
        .where(
            RentalPayment.property_id == prop.id,
            extract("year", RentalPayment.payment_date) == year,
            _confirmed_filter(),
        )
        .group_by("month", "category")
        .order_by("month", "category")
    )
    income_data = income_rows.all()

    # --- Expenses by category per month ---
    expense_rows = await db.execute(
        select(
            extract("month", Expense.date).label("month"),
            Expense.category.label("category"),
            func.sum(Expense.amount).label("total"),
        )
        .where(
            Expense.property_id == prop.id,
            extract("year", Expense.date) == year,
            _confirmed_expense_filter(),
        )
        .group_by("month", "category")
        .order_by("month", "category")
    )
    expense_data = expense_rows.all()

    # Collect all unique category names
    income_categories = sorted({row.category or "Rent" for row in income_data})
    expense_categories = sorted({row.category for row in expense_data})

    # Build lookup: {month: {category: amount}}
    income_by_month: dict[int, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in income_data:
        income_by_month[int(row.month)][row.category or "Rent"] = float(row.total)

    expense_by_month: dict[int, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in expense_data:
        expense_by_month[int(row.month)][row.category] = float(row.total)

    # Build monthly summaries
    monthly_summaries = []
    total_rent = 0.0
    total_expenses = 0.0

    for month in range(1, 13):
        month_income = income_by_month[month]
        month_expense = expense_by_month[month]

        income_items = [
            CategoryMonthlySummary(category=cat, amount=month_income.get(cat, 0))
            for cat in income_categories
        ]
        expense_items = [
            CategoryMonthlySummary(category=cat, amount=month_expense.get(cat, 0))
            for cat in expense_categories
        ]

        rent_total = sum(month_income.values())
        exp_total = sum(month_expense.values())
        total_rent += rent_total
        total_expenses += exp_total

        monthly_summaries.append(
            MonthlySummary(
                month=month,
                rent_collected=rent_total,
                total_expenses=exp_total,
                income_by_category=income_items,
                expense_by_category=expense_items,
            )
        )

    return PropertyYearSummary(
        property_id=prop.id,
        property_name=prop.name,
        monthly_summaries=monthly_summaries,
        total_rent=total_rent,
        total_expenses=total_expenses,
        income_categories=income_categories,
        expense_categories=expense_categories,
    )


async def _generate_excel(
    summaries: list[PropertyYearSummary],
    year: int,
) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=14)
    col_header_font = Font(bold=True, size=10, color="FFFFFF")
    total_font = Font(bold=True, size=11)
    currency_fmt = '#,##0.00'
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    income_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    expense_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    total_income_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    total_expense_fill = PatternFill(start_color="EF5350", end_color="EF5350", fill_type="solid")
    net_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    month_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    total_row_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for idx, ps in enumerate(summaries):
        if idx == 0:
            ws = wb.active
            ws.title = ps.property_name[:31]
        else:
            ws = wb.create_sheet(title=ps.property_name[:31])

        # Build column layout:
        # Col 1: Month
        # Income category columns...
        # Total Income column
        # Expense category columns...
        # Total Expenses column
        # Net Income column
        inc_cats = ps.income_categories
        exp_cats = ps.expense_categories

        col = 1
        col_month = col  # 1

        col_inc_start = 2
        col_inc_end = col_inc_start + len(inc_cats) - 1
        col_total_inc = col_inc_end + 1

        col_exp_start = col_total_inc + 1
        col_exp_end = col_exp_start + len(exp_cats) - 1
        col_total_exp = col_exp_end + 1

        col_net = col_total_exp + 1
        total_cols = col_net

        row = 1
        # Title
        ws.cell(row=row, column=1, value=f"Year-End Report {year} — {ps.property_name}").font = header_font
        row += 2

        # === Column headers ===
        c = ws.cell(row=row, column=col_month, value="Month")
        c.font = Font(bold=True, size=10)
        c.fill = month_fill

        # Income category headers
        for i, cat in enumerate(inc_cats):
            c = ws.cell(row=row, column=col_inc_start + i, value=cat)
            c.font = col_header_font
            c.fill = income_fill
            c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=row, column=col_total_inc, value="Total Income")
        c.font = col_header_font
        c.fill = total_income_fill
        c.alignment = Alignment(horizontal="center")

        # Expense category headers
        for i, cat in enumerate(exp_cats):
            c = ws.cell(row=row, column=col_exp_start + i, value=cat)
            c.font = col_header_font
            c.fill = expense_fill
            c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=row, column=col_total_exp, value="Total Expenses")
        c.font = col_header_font
        c.fill = total_expense_fill
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=row, column=col_net, value="Net Income")
        c.font = col_header_font
        c.fill = net_fill
        c.alignment = Alignment(horizontal="center")

        row += 1

        # === Monthly rows (vertical) ===
        annual_totals = [0.0] * (total_cols + 1)  # 1-based indexing

        for m in range(12):
            ms = ps.monthly_summaries[m]

            # Month name
            c = ws.cell(row=row, column=col_month, value=MONTH_NAMES[m])
            c.font = Font(bold=True, size=10)
            c.fill = month_fill

            # Income categories
            for i, cat in enumerate(inc_cats):
                amt = next((item.amount for item in ms.income_by_category if item.category == cat), 0)
                val = float(amt)
                c = ws.cell(row=row, column=col_inc_start + i, value=val)
                c.number_format = currency_fmt
                c.alignment = Alignment(horizontal="right")
                annual_totals[col_inc_start + i] = annual_totals[col_inc_start + i] + val

            # Total income
            inc_total = float(ms.rent_collected)
            c = ws.cell(row=row, column=col_total_inc, value=inc_total)
            c.number_format = currency_fmt
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="right")
            annual_totals[col_total_inc] = annual_totals[col_total_inc] + inc_total

            # Expense categories
            for i, cat in enumerate(exp_cats):
                amt = next((item.amount for item in ms.expense_by_category if item.category == cat), 0)
                val = float(amt)
                c = ws.cell(row=row, column=col_exp_start + i, value=val)
                c.number_format = currency_fmt
                c.alignment = Alignment(horizontal="right")
                annual_totals[col_exp_start + i] = annual_totals[col_exp_start + i] + val

            # Total expenses
            exp_total = float(ms.total_expenses)
            c = ws.cell(row=row, column=col_total_exp, value=exp_total)
            c.number_format = currency_fmt
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="right")
            annual_totals[col_total_exp] = annual_totals[col_total_exp] + exp_total

            # Net income
            net = inc_total - exp_total
            c = ws.cell(row=row, column=col_net, value=net)
            c.number_format = currency_fmt
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(horizontal="right")
            if net < 0:
                c.font = Font(bold=True, size=10, color="C62828")
            annual_totals[col_net] = annual_totals[col_net] + net

            # Bottom border for the row
            for cc in range(1, total_cols + 1):
                ws.cell(row=row, column=cc).border = thin_border

            row += 1

        # === TOTAL row ===
        row += 1
        c = ws.cell(row=row, column=col_month, value="ANNUAL TOTAL")
        c.font = total_font
        c.fill = total_row_fill

        for cc in range(col_inc_start, total_cols + 1):
            val = annual_totals[cc]
            c = ws.cell(row=row, column=cc, value=val)
            c.number_format = currency_fmt
            c.font = total_font
            c.alignment = Alignment(horizontal="right")
            c.fill = total_row_fill
            if cc == col_net and val < 0:
                c.font = Font(bold=True, size=11, color="C62828")

        # Auto-fit column widths
        ws.column_dimensions[get_column_letter(col_month)].width = 16
        for cc in range(col_inc_start, total_cols + 1):
            ws.column_dimensions[get_column_letter(cc)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("/reports/year-end/{year}")
async def year_end_report_all(
    year: int,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_user),
):
    stmt = select(Property).where(Property.is_active == True)  # noqa: E712
    accessible = await get_accessible_property_ids(user, db)
    if accessible is not None:
        if not accessible:
            raise HTTPException(
                status_code=404, detail="No accessible properties"
            )
        stmt = stmt.where(Property.id.in_(accessible))
    result = await db.execute(stmt)
    properties = result.scalars().all()

    if not properties:
        raise HTTPException(
            status_code=404, detail="No active properties found"
        )

    summaries = []
    for prop in properties:
        summary = await _build_property_summary(db, prop, year)
        summaries.append(summary)

    output = await _generate_excel(summaries, year)

    return StreamingResponse(
        output,
        media_type=EXCEL_CONTENT_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="year_end_report_{year}.xlsx"'
            )
        },
    )


@router.get("/reports/year-end/{year}/{property_id}")
async def year_end_report_property(
    year: int,
    property_id: UUID,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_user),
):
    await require_property_access(property_id, user, db)
    prop = await db.get(Property, property_id)
    if not prop:
        raise HTTPException(status_code=404, detail="Property not found")

    summary = await _build_property_summary(db, prop, year)
    output = await _generate_excel([summary], year)

    return StreamingResponse(
        output,
        media_type=EXCEL_CONTENT_TYPE,
        headers={
            "Content-Disposition": (
                f'attachment; filename="year_end_report_{year}_{prop.name}.xlsx"'
            )
        },
    )


@router.get("/reports/summary/{year}", response_model=YearEndSummary)
async def year_summary(
    year: int,
    db: AsyncSession = Depends(get_db),
    user=Depends(require_user),
):
    stmt = select(Property).where(Property.is_active == True)  # noqa: E712
    accessible = await get_accessible_property_ids(user, db)
    if accessible is not None:
        if not accessible:
            return YearEndSummary(year=year, properties=[])
        stmt = stmt.where(Property.id.in_(accessible))
    result = await db.execute(stmt)
    properties = result.scalars().all()

    summaries = []
    for prop in properties:
        summary = await _build_property_summary(db, prop, year)
        summaries.append(summary)

    return YearEndSummary(year=year, properties=summaries)
